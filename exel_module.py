import pandas as pd

from openpyxl import load_workbook, Workbook


def exel_read():

    try:
        exel_file = load_workbook(filename = 'storage\exel_handbook.xlsx')
        title = exel_file['Handbook']
        dict_reading = {'name': [val.value for i, val in enumerate(title['A']) if i!=0],
                        'number': [val.value for i, val in enumerate(title['B']) if i!=0],
                        'city': [val.value for i, val in enumerate(title['C']) if i!=0]}
        read_handbook = pd.DataFrame(dict_reading)
        return read_handbook
        
    except FileNotFoundError:
        new_handbook = pd.DataFrame({'name':[], 'number':[], 'city':[]})
        new_handbook.index.name = '№'
        return new_handbook


def exel_write(new_handbook):

    exel_file = Workbook()
    exel_list = exel_file.active
    exel_list.title = 'Handbook'
    dict_writing = {'A': [val for val in new_handbook['name']],    
                    'B': [val for val in new_handbook['number']], 
                    'C': [val for val in new_handbook['city']]}

    for key in dict_writing:
        for i, val in enumerate(dict_writing[key],2):
            exel_list[f"{key}{i}"] = val
    else:
        exel_list['A1'] = 'name'
        exel_list['B1'] = 'number'
        exel_list['C1'] = 'city'

    exel_file.save(filename = 'storage\exel_handbook.xlsx')

